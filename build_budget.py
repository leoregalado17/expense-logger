#!/usr/bin/env python3
"""Generate budget_2026.xlsx matching the budget_2026new.xlsx layout.

Structure:
  - Summary: Annual view — months as rows, categories as columns
  - Jan 2026 … Dec 2026: all 7 categories side-by-side (21 cols), 50 data rows
  - VBA Setup: auto-date macro instructions
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

MONTHS = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

CATEGORIES = [
    ("🍽️", "Eating Out"),
    ("🍻", "Drinking"),
    ("🛍️", "Shopping (want)"),
    ("🛒", "Shopping (need)"),
    ("🥦", "Groceries"),
    ("✈️", "Travel - Flights & Hotel"),
    ("🍜", "Travel - Eat & Drink"),
]

# Per-category accent colors for the date column (matches budget_2026new.xlsx exactly)
CAT_COLORS = ["7C6AF7", "56CFB2", "F7846A", "4488FF", "F7C56A", "CF56CF", "6AB8F7"]

# Per-category colors used in the Summary totals row
ANNUAL_COLORS = ["7C6AF7", "56CFB2", "F7846A", "4488FF", "F7C56A", "CF56CF", "6AB8F7"]

# ── Fills ──
BG         = PatternFill("solid", fgColor="1E1E2E")
ROW_EVEN   = PatternFill("solid", fgColor="252535")
ROW_ODD    = PatternFill("solid", fgColor="1E1E2C")
HEADER_BG  = PatternFill("solid", fgColor="2A2A3E")

# ── Shared fonts ──
DESC_FONT  = Font(name="Calibri", color="B0B0C8", size=10)
AMT_FONT   = Font(name="Calibri", color="4488FF", size=10)
HDR_FONT   = Font(name="Calibri", color="7070A0", size=8)
CAT_HDR    = Font(name="Calibri", color="FFFFFF", size=9,  bold=True)
SUBH_WHITE = Font(name="Calibri", color="FFFFFF", size=10, bold=True)

MONEY_FMT  = '$#,##0.00;($#,##0.00);"-"'
DATE_FMT   = "MMM DD"


def f(color, size=10, bold=False):
    return Font(name="Calibri", color=color, size=size, bold=bold)


def fill_bg(ws, max_row, max_col):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).fill = BG


def build_month_sheet(wb, month_abbr):
    name = f"{month_abbr} 2026"
    ws = wb.create_sheet(title=name)
    fill_bg(ws, 60, 22)

    # ── Row 1: Title ──
    ws.merge_cells("A1:U1")
    ws["A1"].value = f"💰  {name}  —  Expense Log"
    ws["A1"].font = f("7C6AF7", 16, True)
    ws["A1"].fill = BG
    ws.row_dimensions[1].height = 34

    # ── Row 2: Category headers (merged over 3 cols each) ──
    ws.row_dimensions[2].height = 28
    col_letters = ["A", "D", "G", "J", "M", "P", "S"]
    merge_ranges = ["A2:C2", "D2:F2", "G2:I2", "J2:L2", "M2:O2", "P2:R2", "S2:U2"]
    for i, (emoji, label) in enumerate(CATEGORIES):
        ws.merge_cells(merge_ranges[i])
        start_col = 1 + i * 3
        cell = ws.cell(row=2, column=start_col)
        cell.value = f"{emoji} {label}"
        cell.font = CAT_HDR
        cell.fill = HEADER_BG
        cell.alignment = Alignment(vertical="center")
        # Fill the merged cells too
        for c in range(start_col, start_col + 3):
            ws.cell(row=2, column=c).fill = HEADER_BG

    # ── Row 3: Sub-headers (Date / Description / Amount) ──
    ws.row_dimensions[3].height = 22
    for i in range(7):
        base = 1 + i * 3
        for j, hdr in enumerate(["Date", "Description", "Amount"]):
            cell = ws.cell(row=3, column=base + j)
            cell.value = hdr
            cell.font = HDR_FONT
            cell.fill = HEADER_BG

    # ── Rows 4–53: Data rows (50 rows) ──
    for r in range(4, 54):
        row_fill = ROW_EVEN if r % 2 == 0 else ROW_ODD
        ws.row_dimensions[r].height = 18
        for i, color in enumerate(CAT_COLORS):
            base = 1 + i * 3
            # Date cell
            dc = ws.cell(r, base)
            dc.fill = row_fill
            dc.font = f(color)
            dc.number_format = DATE_FMT
            # Description cell
            wc = ws.cell(r, base + 1)
            wc.fill = row_fill
            wc.font = DESC_FONT
            # Amount cell
            ac = ws.cell(r, base + 2)
            ac.fill = row_fill
            ac.font = AMT_FONT
            ac.number_format = MONEY_FMT

    # ── Row 54: Category totals ──
    total_colors = ["7C6AF7", "56CFB2", "F7846A", "4488FF", "F7C56A", "CF56CF", "6AB8F7"]
    amt_cols = [3, 6, 9, 12, 15, 18, 21]  # C, F, I, L, O, R, U
    for i, (color, amt_col) in enumerate(zip(total_colors, amt_cols)):
        lbl_col = amt_col - 1
        lbl = ws.cell(54, lbl_col)
        lbl.value = "TOTAL"
        lbl.font = f(color, 10, True)
        lbl.fill = HEADER_BG

        tot = ws.cell(54, amt_col)
        tot.value = f"=SUM({tot.column_letter}4:{tot.column_letter}53)"
        tot.font = f(color, 10, True)
        tot.fill = HEADER_BG
        tot.number_format = MONEY_FMT

    # ── Row 55: Month total ──
    ws.merge_cells("A55:T55")
    mt_label = ws.cell(55, 1)
    mt_label.value = "MONTH TOTAL"
    mt_label.font = f("FFFFFF", 10, True)
    mt_label.fill = HEADER_BG
    mt_total = ws.cell(55, 21)   # U55
    mt_total.value = "=C54+F54+I54+L54+O54+R54+U54"
    mt_total.font = f("7C6AF7", 10, True)
    mt_total.fill = HEADER_BG
    mt_total.number_format = MONEY_FMT

    # ── Column widths ──
    for i in range(7):
        base_letter = chr(ord('A') + i * 3)
        desc_letter = chr(ord('A') + i * 3 + 1)
        amt_letter  = chr(ord('A') + i * 3 + 2)
        ws.column_dimensions[base_letter].width = 9
        ws.column_dimensions[desc_letter].width = 18
        ws.column_dimensions[amt_letter].width  = 10


def build_summary(wb):
    ws = wb.create_sheet(title="Summary")
    fill_bg(ws, 20, 10)

    # Row 1: Title
    ws.merge_cells("A1:I1")
    ws["A1"].value = "💰  Annual Summary"
    ws["A1"].font = f("56CFB2", 18, True)
    ws["A1"].fill = BG
    ws.row_dimensions[1].height = 36

    # Row 2: Subtitle
    ws.merge_cells("A2:I2")
    ws["A2"].value = "Spending by month and category — totals pull live from each month tab"
    ws["A2"].font = f("7070A0", 9)
    ws["A2"].fill = BG
    ws.row_dimensions[2].height = 18

    # Row 3: Column headers
    ws.row_dimensions[3].height = 30
    headers = ["Month"] + [f"{e} {l}" for e, l in CATEGORIES] + ["Month Total"]
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(3, c, hdr)
        cell.font = f("FFFFFF", 10 if c in (1, 9) else 8, True)
        cell.fill = HEADER_BG

    # Category column mapping: C, F, I, L, O, R, U (row 54 on each month sheet)
    cat_sum_cols = ["C", "F", "I", "L", "O", "R", "U"]

    # Rows 4–15: Jan–Dec
    for m_idx, month in enumerate(MONTHS):
        row = 4 + m_idx
        ws.row_dimensions[row].height = 22
        row_fill = ROW_EVEN if row % 2 == 0 else ROW_ODD
        sheet_name = f"{month} 2026"

        ws.cell(row, 1).value = sheet_name
        ws.cell(row, 1).font = f("B0B0C8", 10, True)
        ws.cell(row, 1).fill = row_fill

        for i, (col_letter, color) in enumerate(zip(cat_sum_cols, CAT_COLORS)):
            cell = ws.cell(row, 2 + i)
            cell.value = f"='{sheet_name}'!{col_letter}54"
            cell.font = f(color, 10)
            cell.fill = row_fill
            cell.number_format = MONEY_FMT

        # Month total (U55)
        mt = ws.cell(row, 9)
        mt.value = f"='{sheet_name}'!U55"
        mt.font = f("7C6AF7", 10, True)
        mt.fill = row_fill
        mt.number_format = MONEY_FMT

    # Row 16: Annual totals
    ws.row_dimensions[16].height = 28
    ws.cell(16, 1).value = "ANNUAL"
    ws.cell(16, 1).font = f("7C6AF7", 12, True)
    ws.cell(16, 1).fill = HEADER_BG

    annual_colors = ["7C6AF7", "56CFB2", "F7846A", "4488FF", "F7C56A", "CF56CF", "6AB8F7", "7C6AF7"]
    annual_sizes  = [11, 11, 11, 11, 11, 11, 11, 12]
    for c in range(2, 10):
        col_letter = chr(ord('A') + c - 1)
        cell = ws.cell(16, c)
        cell.value = f"=SUM({col_letter}4:{col_letter}15)"
        cell.font = f(annual_colors[c - 2], annual_sizes[c - 2], True)
        cell.fill = HEADER_BG
        cell.number_format = MONEY_FMT

    # Column widths
    ws.column_dimensions["A"].width = 12
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["I"].width = 14


def build_vba_setup(wb):
    ws = wb.create_sheet(title="VBA Setup")
    fill_bg(ws, 60, 3)

    lines = [
        (1,  "⚡  Enable Auto-Date (one-time setup)",          f("56CFB2", 16, True)),
        (3,  "WINDOWS",                                        f("7C6AF7", 12, True)),
        (4,  "Step 1 → Press  Alt + F11  to open the Visual Basic Editor", f("B0B0C8", 10)),
        (5,  "Step 2 → In the left panel, double-click 'ThisWorkbook'",    f("B0B0C8", 10)),
        (6,  "Step 3 → Copy ALL the code below and paste it",              f("B0B0C8", 10)),
        (7,  "Step 4 → Ctrl + S  (save as .xlsm if prompted) → close VBA editor ✅", f("B0B0C8", 10)),
        (9,  "MAC",                                            f("7C6AF7", 12, True)),
        (10, "Step 1 → Tools > Macros > Edit Macros",         f("B0B0C8", 10)),
        (11, "Step 2 → Double-click 'ThisWorkbook' in the left panel", f("B0B0C8", 10)),
        (12, "Step 3 → Copy ALL the code below, paste, save ✅",       f("B0B0C8", 10)),
        (14, "⚠️  Click 'Enable Macros' when opening the file.",        f("F7846A", 10, True)),
        (16, "━━━  PASTE THIS CODE  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", f("56CFB2", 10)),
    ]
    for row, text, font in lines:
        ws.cell(row, 1).value = text
        ws.cell(row, 1).font = font
        ws.cell(row, 1).fill = BG

    vba = [
        'Private Sub Workbook_SheetChange(ByVal Sh As Object, ByVal Target As Range)',
        '    Dim cell As Range',
        '    If Sh.Name = "Summary" Or Sh.Name = "VBA Setup" Then Exit Sub',
        '    Application.EnableEvents = False',
        '    On Error GoTo Cleanup',
        '    For Each cell In Target',
        "        ' Date col = col 1 (A), Description = col 2 (B), Amount = col 3 (C)",
        "        ' But we have blocks: each category block is 3 cols wide",
        "        ' Col pattern: A=date, B=desc, C=amount, D=date, E=desc, F=amount, ...",
        "        ' Every 3rd col starting at col 1 is a date col",
        "        ' So date cols: 1,4,7,10,13,16,19",
        "        ' Desc cols: 2,5,8,11,14,17,20",
        "        ' Amt cols: 3,6,9,12,15,18,21",
        '        Dim colInBlock As Integer',
        '        colInBlock = ((cell.Column - 1) Mod 3) + 1',
        "        ' colInBlock=1 is date (skip), =2 is desc, =3 is amount",
        '        If colInBlock = 2 Or colInBlock = 3 Then',
        "            ' find the date cell for this block",
        '            Dim dateCol As Integer',
        '            dateCol = cell.Column - colInBlock + 1',
        '            Dim dateCell As Range',
        '            Set dateCell = Sh.Cells(cell.Row, dateCol)',
        '            If cell.Value <> "" And dateCell.Value = "" Then',
        '                dateCell.Value = Date',
        '                dateCell.NumberFormat = "MMM DD"',
        '            End If',
        '            Dim descCell As Range, amtCell As Range',
        '            Set descCell = Sh.Cells(cell.Row, dateCol + 1)',
        '            Set amtCell  = Sh.Cells(cell.Row, dateCol + 2)',
        '            If descCell.Value = "" And amtCell.Value = "" Then',
        '                dateCell.ClearContents',
        '            End If',
        '        End If',
        '    Next cell',
        'Cleanup:',
        '    Application.EnableEvents = True',
        'End Sub',
    ]
    code_font = Font(name="Consolas", color="56CF72", size=10)
    for i, line in enumerate(vba):
        cell = ws.cell(17 + i, 1)
        cell.value = line
        cell.font = code_font
        cell.fill = BG

    ws.column_dimensions["A"].width = 80


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # Build Summary first
    build_summary(wb)

    # Build 12 monthly sheets
    for month in MONTHS:
        build_month_sheet(wb, month)

    # Build VBA Setup last
    build_vba_setup(wb)

    out = "budget_2026.xlsx"
    wb.save(out)
    print(f"✓ Created {out}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Layout: Annual Summary + 12 monthly sheets (50 rows × 7 categories) + VBA Setup")


if __name__ == "__main__":
    main()
